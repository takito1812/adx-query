"""
Output helpers for presenting query results.
"""

from __future__ import annotations

import csv
import io
import json
from typing import Iterable, List, Optional, Sequence, TextIO


def write_json(rows: Iterable[dict], stream: TextIO) -> None:
    first = True
    stream.write("[\n")
    for row in rows:
        if not first:
            stream.write(",\n")
        json.dump(row, stream, ensure_ascii=False)
        first = False
    if not first:
        stream.write("\n")
    stream.write("]")


def write_csv(
    rows: Iterable[dict],
    stream: TextIO,
    fieldnames: Optional[Sequence[str]] = None,
) -> None:
    iterator = iter(rows)
    try:
        first_row = next(iterator)
    except StopIteration:
        if fieldnames is not None:
            writer = csv.DictWriter(stream, fieldnames=fieldnames)
            writer.writeheader()
        return

    if fieldnames is None:
        fieldnames = list(first_row.keys())

    writer = csv.DictWriter(stream, fieldnames=fieldnames)
    writer.writeheader()
    writer.writerow(_stringify_row(first_row, fieldnames))
    for row in iterator:
        writer.writerow(_stringify_row(row, fieldnames))


def write_table(
    rows: Iterable[dict],
    stream: TextIO,
    field_order: Optional[Sequence[str]] = None,
) -> None:
    rows_list = list(rows)
    if not rows_list:
        return

    if field_order is None:
        # preserve insertion order from first row
        field_order = list(rows_list[0].keys())

    widths = {field: len(field) for field in field_order}
    formatted_rows: List[List[str]] = []

    for row in rows_list:
        formatted_row: List[str] = []
        for field in field_order:
            value = row.get(field, "")
            if isinstance(value, (list, tuple)):
                cell = ", ".join(map(str, value))
            else:
                cell = str(value)
            widths[field] = max(widths[field], len(cell))
            formatted_row.append(cell)
        formatted_rows.append(formatted_row)

    header = " | ".join(field.ljust(widths[field]) for field in field_order)
    separator = "-+-".join("-" * widths[field] for field in field_order)
    stream.write(header + "\n")
    stream.write(separator + "\n")
    for formatted_row in formatted_rows:
        line = " | ".join(
            formatted_row[i].ljust(widths[field_order[i]])
            for i in range(len(field_order))
        )
        stream.write(line + "\n")


def _stringify_row(row: dict, fieldnames: Sequence[str]) -> dict:
    flattened = {}
    for field in fieldnames:
        value = row.get(field, "")
        if isinstance(value, (list, tuple)):
            flattened[field] = ";".join(map(str, value))
        else:
            flattened[field] = value
    return flattened


def write_excel(
    rows: Iterable[dict],
    path: str,
    fieldnames: Optional[Sequence[str]] = None,
    table_name: str = "ResultsTable",
) -> None:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    rows_list = list(rows)

    if fieldnames is None:
        if rows_list:
            fieldnames = list(rows_list[0].keys())
        else:
            fieldnames = []

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    if fieldnames:
        ws.append(list(fieldnames))

    for row in rows_list:
        ws.append([_stringify_cell(row.get(field, "")) for field in fieldnames])

    if fieldnames:
        end_col = get_column_letter(len(fieldnames))
        end_row = len(rows_list) + 1
        table = Table(displayName=table_name, ref=f"A1:{end_col}{max(end_row,1)}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
        # Wrap in try/except in case table already exists (shouldn't happen but safe)
        try:
            ws.add_table(table)
        except ValueError:
            pass

        for idx, field in enumerate(fieldnames, start=1):
            column_letter = get_column_letter(idx)
            max_length = len(str(field))
            for cell in ws[column_letter]:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2

    wb.save(path)


def _stringify_cell(value) -> str:
    if isinstance(value, (list, tuple)):
        return ";".join(map(str, value))
    if value is None:
        return ""
    return str(value)
